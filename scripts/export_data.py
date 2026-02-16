"""
Export product and category data from SQLite DB + Excel pricelists into JSON files
for the Next.js web app.

Sources:
- products.db (794 products with LV names, descriptions, categories)
- LAB20 CATALOGUE PRICELIST cenu analīze.xlsx (1,668 products with EN names, prices, EAN)
- laborpro_nextcloud_catalog.xlsx (brand info)

Outputs:
- data/categories.json
- data/products.json
"""

import sqlite3
import json
import os
import sys
import re
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WEBPAGECOPY_DIR = os.path.dirname(BASE_DIR)
DATA_DIR = os.path.join(WEBPAGECOPY_DIR, 'data')
DB_PATH = os.path.join(DATA_DIR, 'products.db')
PRICELIST_PATH = os.path.join(DATA_DIR, 'LAB20 CATALOGUE PRICELIST cenu analīze.xlsx')
NEXTCLOUD_PATH = os.path.join(DATA_DIR, 'laborpro_nextcloud_catalog.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'data')
LOCAL_IMAGES_DIR = os.path.join(BASE_DIR, 'public', 'images', 'products')

# Chapter (pricelist) -> our category number mapping
CHAPTER_TO_CATEGORY = {
    '01': 2,   # Flat and Curling Irons
    '02': 1,   # Hair Dryers
    '03': 3,   # Hair Clippers
    '04': 4,   # Barber and Accessories
    '05': 8,   # Scissors
    '06': 6,   # Combs
    '07': 5,   # Brushes
    '08': 9,   # Colouring and Styling
    '09': 18,  # Hair Extensions and Coiffure
    '10': 10,  # Hair Accessories
    '11': 19,  # Training Heads
    '12': 21,  # Cases and Bags
    '13': 20,  # Apparel and Clothing
    # 14 missing from pricelist -> Disposable Items (category 11), DB only
    '15': 12,  # Sterilization and Hygiene
    '16': 22,  # Furnishings
    '17': 16,  # Hair Removal
    '18': 14,  # Manicure and Nails
    '19': 7,   # Steel Tools
    '20': 15,  # Pedicure
    '21': 17,  # Wellness and Treatment Tools
    '22': 13,  # Make-Up and Eyelashes
}

# Manual SKU -> category overrides for misplaced products
# Format: 'SKU' (uppercase): category_id
SKU_CATEGORY_OVERRIDES = {
    # From Cat 01 (Hair Dryers) — shampoos/conditioners → 09 Colouring and Styling
    '3822': 9,
    '54308': 9,
    'CRL350': 9,
    'SRL350': 9,
    'SRM350': 9,
    'SRR350': 9,
    '45493': 9,
    # From Cat 01 — curling irons/straighteners → 02 Flat and Curling Irons
    '54022': 2,
    'B132 \xa0 \xa0 \xa0B133': 2,  # SKU as stored with nbsp
    'B255TL-1': 2,
    'UG09': 2,
    'UG125': 2,
    'UG156': 2,
    'UG157': 2,
    'UG170': 2,
    # From Cat 01 — clipper → 03 Hair Clippers
    'B507': 3,
    # From Cat 01 — epilator → 16 Hair Removal
    'B805': 16,
    # From Cat 01 — abrasive disc → 15 Pedicure
    'E374': 15,
    # From Cat 02 (Flat and Curling Irons) — mini dryers → 01 Hair Dryers
    'B348 B349 B350 B351': 1,
    # From Cat 02 — thermal brushes → 05 Brushes
    'C820': 5,
    'C821': 5,
    'C822': 5,
    'C830': 5,
    # From Cat 02 — diffuser → 01 Hair Dryers
    'E390': 1,
    # From Cat 03 (Hair Clippers) — split ender → 09 Colouring and Styling
    'B904': 9,
    # From Cat 05 (Brushes) — dry shampoo → 09 Colouring and Styling
    'HC TRAVEL DRY SHAMPOO': 9,
    # From Cat 07 (Steel Tools) — wax dissolver → 16 Hair Removal
    'H306': 16,
    # From Cat 09 (Colouring and Styling) — pedicure discs → 15 Pedicure
    'E341': 15,
    'E350': 15,
    'E353': 15,
    'E354': 15,
    'E355/B': 15,
    'E356': 15,
    'E357': 15,
    'E361': 15,
    'E369': 15,
    # From Cat 09 — flat iron → 02 Flat and Curling Irons
    'E359': 2,
    # From Cat 09 — suction hood → 22 Furnishings
    'B053': 22,
    # From Cat 12 (Sterilization) — pedicure disc → 15 Pedicure
    'E346': 15,
    # From Cat 12 — silk perfume → 09 Colouring and Styling
    'ST TRAVEL SILK PERFUME': 9,
    # From Cat 14 (Manicure) — aluminum roll → 09 Colouring and Styling
    'H140A': 9,
    # From Cat 14 — facial brush → 17 Wellness and Treatment Tools
    'H827': 17,
    # From Cat 17 (Wellness) — eyelash/makeup products → 13 Make-Up and Eyelashes
    '32941': 13,
    'IN007': 13,
    'IN011': 13,
    'IN012': 13,
    'IN013': 13,
    # From Cat 17 — solarium glasses → 20 Apparel and Clothing
    'F730': 20,
    # From Cat 20 (Apparel) — hair dryer → 01 Hair Dryers
    'W042': 1,
    # From Cat 22 (Furnishings) — wipes → 09 Colouring and Styling
    '59208': 9,
    # From Cat 22 — hair dryer → 01 Hair Dryers
    '59231': 1,
    # From Cat 22 — nail buffer → 14 Manicure and Nails
    'E336': 14,
    # From Cat 22 — disposable aprons → 11 Disposable Items
    'G003': 11,
}


def load_categories_from_db():
    """Load all 22 categories from the database."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT id, name_en, name_lv, slug, sort_order FROM categories ORDER BY sort_order')
    categories = {}
    for row in c.fetchall():
        categories[row[0]] = {
            'id': row[0],
            'number': f'{row[4]:02d}',
            'name_en': row[1],
            'name_lv': row[2],
            'slug': row[3],
            'productCount': 0,
        }
    conn.close()
    return categories


def load_db_products():
    """Load products from SQLite with their images."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Get products
    c.execute('''
        SELECT p.id, p.sku, p.name_lv, p.description_lv, p.price, p.category_id
        FROM products p
        ORDER BY p.sku
    ''')
    products = {}
    for row in c.fetchall():
        sku = row[1]
        if sku:
            products[sku.strip().upper()] = {
                'sku': row[1],
                'name_lv': row[2],
                'description_lv': row[3],
                'price_db': row[4],
                'category_id': row[5],
            }

    # Get images
    c.execute('''
        SELECT p.sku, pi.image_url, pi.sort_order
        FROM product_images pi
        JOIN products p ON pi.product_id = p.id
        ORDER BY p.sku, pi.sort_order
    ''')
    for row in c.fetchall():
        sku = row[0].strip().upper() if row[0] else None
        if sku and sku in products:
            if 'images' not in products[sku]:
                products[sku]['images'] = []
            products[sku]['images'].append(row[1])

    conn.close()
    return products


def load_pricelist():
    """Load products from the pricelist Excel."""
    wb = openpyxl.load_workbook(PRICELIST_PATH, read_only=True, data_only=True)
    ws = wb['Verifica']

    products = []
    for row in ws.iter_rows(min_row=2, max_col=14, values_only=True):
        sku = row[0]
        if not sku:
            continue
        sku = str(sku).strip()
        chapter = str(row[3]).strip() if row[3] else None

        name_lv = row[1] if row[1] and row[1] != 'Nav atrasts' else None
        name_en = row[2] if row[2] else None
        price = row[5] if row[5] else None  # CATALOGUE PRICE (without VAT)
        ean = str(row[7]).strip() if row[7] else None

        # Map chapter to our category
        category_id = CHAPTER_TO_CATEGORY.get(chapter) if chapter else None

        products.append({
            'sku': sku,
            'name_lv_pricelist': name_lv,
            'name_en': name_en,
            'price': float(price) if price else None,
            'ean': ean,
            'chapter': chapter,
            'category_id': category_id,
        })

    wb.close()
    return products


def load_nextcloud_brands():
    """Load brand info from nextcloud catalog."""
    wb = openpyxl.load_workbook(NEXTCLOUD_PATH, read_only=True, data_only=True)
    ws = wb['Catalog']

    brands = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        sku = row[0]
        brand = row[1]
        if sku and brand:
            brands[str(sku).strip().upper()] = str(brand).strip()

    wb.close()
    return brands


def get_local_images(sku):
    """Find local images for a product SKU in public/images/products/{SKU}/."""
    sku_dir = os.path.join(LOCAL_IMAGES_DIR, sku)
    if not os.path.isdir(sku_dir):
        return []
    image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
    images = []
    for f in sorted(os.listdir(sku_dir)):
        ext = os.path.splitext(f)[1].lower()
        if ext in image_exts:
            images.append(f'/images/products/{sku}/{f}')
    return images


def slugify(text):
    """Create a URL-safe slug from text."""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')


def main():
    print("Loading data sources...")

    # Load all sources
    categories = load_categories_from_db()
    db_products = load_db_products()
    pricelist_products = load_pricelist()
    brands = load_nextcloud_brands()

    print(f"  DB categories: {len(categories)}")
    print(f"  DB products: {len(db_products)}")
    print(f"  Pricelist products: {len(pricelist_products)}")
    print(f"  Nextcloud brands: {len(brands)}")

    # Build merged product list
    # Start with pricelist as primary (most products, has prices and EN names)
    merged = {}
    product_id = 0

    for pl in pricelist_products:
        sku_upper = pl['sku'].upper()
        db = db_products.get(sku_upper, {})
        brand = brands.get(sku_upper, None)

        # Determine best name_lv: prefer DB, then pricelist
        name_lv = db.get('name_lv') or pl.get('name_lv_pricelist') or pl.get('name_en', '')
        name_en = pl.get('name_en', '')

        # Determine category: override > pricelist mapping > DB
        category_id = SKU_CATEGORY_OVERRIDES.get(pl['sku']) or SKU_CATEGORY_OVERRIDES.get(sku_upper) or pl.get('category_id') or db.get('category_id')
        if not category_id:
            continue  # Skip products without a category

        # Price: prefer pricelist, fallback DB
        price = pl.get('price') or db.get('price_db')

        product_id += 1
        category_slug = categories[category_id]['slug'] if category_id in categories else ''

        # Prefer local Nextcloud images, fallback to scraped remote URLs
        local_imgs = get_local_images(pl['sku'])
        images = local_imgs if local_imgs else db.get('images', [])

        merged[sku_upper] = {
            'id': product_id,
            'sku': pl['sku'],
            'name_lv': name_lv,
            'name_en': name_en,
            'description_lv': db.get('description_lv', ''),
            'price': price,
            'categoryId': category_id,
            'categorySlug': category_slug,
            'brand': brand or '',
            'ean': pl.get('ean', ''),
            'images': images,
        }

    # Add DB-only products (e.g., Disposable Items category 11)
    for sku_upper, db in db_products.items():
        if sku_upper not in merged:
            category_id = SKU_CATEGORY_OVERRIDES.get(db['sku']) or SKU_CATEGORY_OVERRIDES.get(sku_upper) or db.get('category_id')
            if not category_id or category_id not in categories:
                continue

            brand = brands.get(sku_upper, None)
            product_id += 1
            category_slug = categories[category_id]['slug']

            local_imgs = get_local_images(db['sku'])
            images = local_imgs if local_imgs else db.get('images', [])

            merged[sku_upper] = {
                'id': product_id,
                'sku': db['sku'],
                'name_lv': db.get('name_lv', ''),
                'name_en': '',
                'description_lv': db.get('description_lv', ''),
                'price': db.get('price_db'),
                'categoryId': category_id,
                'categorySlug': category_slug,
                'brand': brand or '',
                'ean': '',
                'images': images,
            }

    # Count products per category
    for prod in merged.values():
        cat_id = prod['categoryId']
        if cat_id in categories:
            categories[cat_id]['productCount'] += 1

    # Sort products by SKU within each category
    products_list = sorted(merged.values(), key=lambda p: (p['categoryId'], p['sku']))
    # Re-assign sequential IDs
    for i, p in enumerate(products_list, 1):
        p['id'] = i

    # Build categories list
    categories_list = sorted(categories.values(), key=lambda c: int(c['number']))

    # Write output
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(os.path.join(OUTPUT_DIR, 'categories.json'), 'w', encoding='utf-8') as f:
        json.dump(categories_list, f, ensure_ascii=False, indent=2)

    with open(os.path.join(OUTPUT_DIR, 'products.json'), 'w', encoding='utf-8') as f:
        json.dump(products_list, f, ensure_ascii=False, indent=2)

    print(f"\nExported:")
    print(f"  {len(categories_list)} categories -> data/categories.json")
    print(f"  {len(products_list)} products -> data/products.json")

    # Summary per category
    print("\nProducts per category:")
    for cat in categories_list:
        print(f"  {cat['number']} {cat['name_lv']}: {cat['productCount']}")


if __name__ == '__main__':
    main()
